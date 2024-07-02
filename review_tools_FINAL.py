import hashlib
import os
import random
import subprocess
import sys
from configparser import ConfigParser

import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QTabWidget, QTableWidget, QTableWidgetItem, \
    QProgressBar, QLabel, QHBoxLayout, QGroupBox, QMainWindow, QLineEdit, QComboBox, QSizePolicy, QMessageBox, \
    QFormLayout, QFileDialog, QDialog, QHeaderView
from PyQt5.QtGui import QPainter, QColor, QBrush, QFont, QPen
from PyQt5.QtCore import Qt, QRect, QTimer, QDateTime
from openpyxl.reader.excel import load_workbook

ENABILITY_TYPE = [
    '画面', 'バッチ', 'API', '共通部品'
]

ENABILITY_SYSTEM = [
    'EnabilityCIS',
    'EnabilityOrder',
    'EnabilityPortal',
    'EnabilityPortal2'
]

INPUT_ID, \
    INPUT_TYPE, \
    INPUT_SYSTEM, \
    INPUT_SVN, \
    INPUT_MANUAL, \
    INPUT_OLD, \
    INPUT_NEW, \
    INPUT_WBS = "", "", "", "", "", "", "", ""

EXCEL_TEST = "単体テスト仕様書"
EXCEL_EVIDENCE = "単体テストエビデンス"
EXCEL_LIST = "成果物一覧"
EXCEL_COMPARE = "手修正確認結果"
EXCEL_COVERAGE = "カバレッジ結果"
EXCEL_REVIEW = "レビュー記録表"
EXCEL_CD_CHECKLIST = "CDチェックリスト"
EXCEL_UT_CHECKLIST = "UTチェックリスト"

variables = {
    'EXCEL_TEST': EXCEL_TEST,
    'EXCEL_EVIDENCE': EXCEL_EVIDENCE,
    'EXCEL_LIST': EXCEL_LIST,
    'EXCEL_COMPARE': EXCEL_COMPARE,
    'EXCEL_COVERAGE': EXCEL_COVERAGE,
    'EXCEL_REVIEW': EXCEL_REVIEW,
    'EXCEL_CD_CHECKLIST': EXCEL_CD_CHECKLIST,
    'EXCEL_UT_CHECKLIST': EXCEL_UT_CHECKLIST
}

EXCEL_TOTAL_MAP = {"EXCEL_TEST": EXCEL_TEST, "EXCEL_EVIDENCE": EXCEL_EVIDENCE, "EXCEL_LIST": EXCEL_LIST,
                   "EXCEL_COMPARE": EXCEL_COMPARE, "EXCEL_COVERAGE": EXCEL_COVERAGE, "EXCEL_REVIEW": EXCEL_REVIEW,
                   "EXCEL_CD_CHECKLIST": EXCEL_CD_CHECKLIST, "EXCEL_UT_CHECKLIST": EXCEL_UT_CHECKLIST}

COMMON_FUNC_ID = ""
COMMON_FUNC_NAME = ""
COMMON_SYSTEM = ""
COMMON_TYPE = ""
COMMON_USER_NAME = ""
COMMON_USER_DATE = ""

WBS_NAME = ""
WBS_DATE = ""


# EXCEL_TOTAL_MAP = [EXCEL_TEST, EXCEL_EVIDENCE, EXCEL_LIST, EXCEL_COMPARE,
#                     EXCEL_COVERAGE, EXCEL_REVIEW, EXCEL_CD_CHECKLIST, EXCEL_UT_CHECKLIST]


def get_program_path():
    """アプリのパスを取得"""
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def get_config_file_path():
    """コンフィグのパスを取得"""
    return os.path.join(get_program_path(), ".review_config.ini")


def init_config_content():
    global INPUT_ID, INPUT_TYPE, INPUT_SYSTEM, INPUT_SVN, INPUT_MANUAL, INPUT_OLD, INPUT_NEW, INPUT_WBS
    try:
        ids = load_config_content('Ids')
        paths = load_config_content('Paths')
    except Exception as e:
        set_message_box("CRITICAL", "コンフィグ", "コンフィグファイルが存在しませんが、チェックしてください。")
        return
    if ids['input_id'] is not None:
        INPUT_ID = ids['input_id']
    if ids['input_type'] is not None:
        INPUT_TYPE = ids['input_type']
    if ids['input_system'] is not None:
        INPUT_SYSTEM = ids['input_system']
    if paths['input_svn'] is not None:
        INPUT_SVN = paths['input_svn']
    if paths['input_manual'] is not None:
        INPUT_MANUAL = paths['input_manual']
    if paths['input_old'] is not None:
        INPUT_OLD = paths['input_old']
    if paths['input_new'] is not None:
        INPUT_NEW = paths['input_new']
    if paths['input_wbs'] is not None:
        INPUT_WBS = paths['input_wbs']


def load_config_content(tag):
    """コンフィグをロード"""
    config = ConfigParser()
    config_path = get_config_file_path()
    if os.path.exists(config_path) is False:
        raise
    config.read(config_path, encoding='utf-8')
    return config[tag] if tag in config else {}


def save_file_paths(self):
    """コンフィグにパスインフォを保存"""
    config = ConfigParser()
    config_path = get_config_file_path()
    if os.path.exists(config_path) is False:
        return
    config.read(config_path, encoding='utf-8')
    if self.parent.input_id.text():
        config.set('Ids', 'input_id', self.parent.input_id.text())
    if self.parent.input_type.currentText():
        config.set('Ids', 'input_type', self.parent.input_type.currentText())
    if self.parent.input_system.currentText():
        config.set('Ids', 'input_system', self.parent.input_system.currentText())
    if self.parent.input_svn.text():
        config.set('Paths', 'input_svn', self.parent.input_svn.text())
    if self.parent.input_manual.text():
        config.set('Paths', 'input_manual', self.parent.input_manual.text())
    if self.parent.input_old.text():
        config.set('Paths', 'input_old', self.parent.input_old.text())
    if self.parent.input_new.text():
        config.set('Paths', 'input_new', self.parent.input_new.text())
    if self.parent.input_wbs.text():
        config.set('Paths', 'input_wbs', self.parent.input_wbs.text())
    with open(get_config_file_path(), 'w', encoding='utf-8') as configfile:
        config.write(configfile)


def set_message_box(message_type, title, context):
    """メッセージを反映"""
    if message_type == 'WARNING':
        QMessageBox.warning(None, title, context)
    if message_type == 'CRITICAL':
        QMessageBox.critical(None, title, context)
    if message_type == 'INFO':
        QMessageBox.information(None, title, context)
    if message_type == 'QUESTION':
        QMessageBox.question(None, title, context)


def svn_version_check():
    print('svn_version_check start...')
    # cmd_update = 'svn update ' + self.parent.input_browse.text()
    # result = os.system(cmd_update)
    # print("svn update result : ", result)
    result = subprocess.run(['svn', '--version'], text=True, capture_output=True, check=False
                            , shell=True, creationflags=subprocess.CREATE_NO_WINDOW)
    if "svn, version" in result.stdout:
        pass
    else:
        set_message_box("WARNING", "SVN",
                        "コンピューターにはまだSVNコマンドラインがインストールされていません。\nインストールしてください。")


def svn_check_file(file_path):
    """SVNチェック"""
    print('svn_check_file start...')
    # "\"" + file_path + "\""
    result = subprocess.run(['svn', 'status', file_path], text=True, capture_output=True, check=False
                            , shell=True, creationflags=subprocess.CREATE_NO_WINDOW)
    if result.returncode == 0:
        version_result = svn_check_file_version(file_path)
        if version_result is True:
            return "SVNに問題ない。"
        else:
            return "本地のバージョンは最新のではありませんので、チェックしてください。"
    else:
        return "SVNにコミットするかどうかのことをチェックしてください。"


def svn_check_file_version(file_path):
    """SVNバージョンチェック"""
    print('svn_check_file_version start...')
    # "\"" + file_path + "\""
    result = subprocess.run(['svn', 'status', '-u', file_path], text=True, capture_output=True, check=False
                            , shell=True, creationflags=subprocess.CREATE_NO_WINDOW)
    if result.returncode == 0:
        return True
    else:
        return False


def svn_operate(self, folder):
    """SVNから更新"""
    print('svn_operate start...')
    # cmd_update = 'svn update ' + self.parent.input_browse.text()
    # result = os.system(cmd_update)
    # print("svn update result : ", result)
    result = subprocess.run(['svn', '--version'], text=True, capture_output=True, check=False
                            , shell=True, creationflags=subprocess.CREATE_NO_WINDOW)
    if "svn, version" in result.stdout:
        with subprocess.Popen(['svn', 'update', folder],
                              stdin=subprocess.PIPE,
                              stdout=subprocess.PIPE,
                              stderr=subprocess.PIPE, text=True) as proc:
            stdout, stderr = proc.communicate()
        if proc.returncode != 0:
            # set_message_box("CRITICAL", "SVN",
            #                 "サンプルフォルダをSVNから最新版に更新することが失敗しました、\n自分で更新してください。")
            print(f"Command '{self.parent.input_browse.text()}' "
                  f"failed with return code {proc.returncode}")
            print("Errors:", stderr)
        else:
            # set_message_box("INFO", "SVN", "「" + folder.split("\\")[len(folder.split("\\")) - 1] + "」" +
            #                 "\nSVNから更新することが成功しました、\n続けてください。")
            print(f"Command '{self.parent.input_browse.text()}' executed successfully")
            print("Output:", stdout)
    else:
        set_message_box("WARNING", "SVN",
                        "コンピューターにはまだSVNコマンドラインがインストールされていません。\nインストールしてください。")


def file_md5(filename):
    hash_md5 = hashlib.md5()
    with open(filename, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()


def files_are_identical(file1, file2):
    return file_md5(file1) == file_md5(file2)


def is_null_check(self):
    """非空チェック"""
    check_flag = False
    context = ""
    if self.parent.input_id.text() is None or self.parent.input_id.text() == '':
        check_flag = True
        context = context + "機能ID\n"
    if self.parent.input_type.currentText() is None or self.parent.input_type.currentText() == '':
        check_flag = True
        context = context + "区分\n"
    if self.parent.input_system.currentText() is None or self.parent.input_system.currentText() == '':
        check_flag = True
        context = context + "システム\n"
    if self.parent.input_svn.text() is None or self.parent.input_svn.text() == '':
        check_flag = True
        context = context + "SVNパス\n"
    if self.parent.input_manual.text() is None or self.parent.input_manual.text() == '':
        check_flag = True
        context = context + "マニュアルパス\n"
    if self.parent.input_old.text() is None or self.parent.input_old.text() == '':
        check_flag = True
        context = context + "元ソース\n"
    if self.parent.input_new.text() is None or self.parent.input_new.text() == '':
        check_flag = True
        context = context + "新ソース\n"
    if self.parent.input_wbs.text() is None or self.parent.input_wbs.text() == '':
        check_flag = True
        context = context + "WBSパス\n"
    return check_flag, context


def find_files(folder_path_to_find, word, suffix):
    print("find_files :" + suffix)
    for root, dirs, files in os.walk(folder_path_to_find):
        for file_con in files:
            if file_con.find(word) > 0 and file_con.find(suffix) > 0:
                if (word.find("Custom_") > 0) == (file_con.find("Custom_") > 0):
                    return os.path.join(root, file_con)
                else:
                    continue
    return None


def find_files_in_svn(svn_path, func_id, suffix, flag):
    excel_path = find_files(svn_path, func_id, suffix)
    if excel_path is None and flag is True:
        set_message_box("CRITICAL", "ファイル", "機能ID「" + func_id + "」の成果物が見つかりませんので、チェックしてください。")
        return None
    else:
        return excel_path


def read_name_excel(path, sheet):
    sw = load_workbook(f'{path}', data_only=True)

    try:
        src_sheet = sw[f'{sheet}']
    except KeyError:
        raise KeyError('シートが存在しません。')

    source_list = []
    for row in src_sheet.iter_rows():
        source_list_list = []
        # if row[2].value != "共通部品":
        #     continue
        for cell in row:
            source_list_list.append(cell.value)
        source_list.append(source_list_list)
    return source_list


def column_letter_to_number(column_letter):
    """ディジットに変更する"""
    column_number = 0
    for char in column_letter:
        column_number = column_number * 26 + (ord(char) - ord("A") + 1)
    return column_number


def set_status_label(self, context):
    self.parent.status_label.setText(context)


def clearAllTables(self):
    for index in range(self.parent.tabs.count()):
        tabWidget = self.parent.tabs.widget(index)
        if tabWidget is not None:
            # 找到该标签页中的 QTableWidget
            tableWidget = tabWidget.findChild(QTableWidget)
            if tableWidget is not None:
                # 清空 QTableWidget 的所有行和数据
                tableWidget.setRowCount(0)
                for row in range(tableWidget.rowCount()):
                    for col in range(tableWidget.columnCount()):
                        item = tableWidget.item(row, col)
                        if item is not None:
                            item.setText("")


class StepProgressBar(QWidget):
    def __init__(self, steps=5):
        super().__init__()
        self.steps = steps
        self.current_step = 0
        self.setMinimumHeight(40)
        self.setMinimumWidth(300)
        self.setMouseTracking(True)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        width = self.width()
        height = self.height()
        step_width = width // self.steps
        radius = 8  # Radius for rounded corners

        for i in range(self.steps):
            if i < self.current_step:
                brush_color = QColor(100, 149, 237)  # Cornflower Blue for completed steps
            else:
                brush_color = QColor(211, 211, 211)  # Light Grey for remaining steps

            pen = QPen(Qt.NoPen)
            painter.setPen(pen)
            painter.setBrush(QBrush(brush_color))

            # Draw rounded rectangle for steps
            painter.drawRoundedRect(QRect(i * step_width, 0, step_width - 5, height), radius, radius)

            # Draw text
            painter.setPen(QColor(255, 255, 255))  # White for text
            font = QFont("Arial", 12, QFont.Bold)
            painter.setFont(font)
            painter.drawText(QRect(i * step_width, 0, step_width - 5, height), Qt.AlignCenter, f"{i + 1}")

    def mousePressEvent(self, event):
        width = self.width()
        step_width = width // self.steps
        clicked_step = event.x() // step_width
        if clicked_step < self.steps:
            self.current_step = clicked_step + 1
            self.update()
            self.parent().tab_switched(clicked_step)

    def advance_step(self):
        if self.current_step < self.steps:
            self.current_step += 1
            self.update()
            self.parent().tab_switched(self.current_step - 1)


class SortableTable(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setColumnCount(4)
        self.setHorizontalHeaderLabels(['区分', 'ファイル', '備考', '状態'])
        self.setSortingEnabled(False)  # Disable default sorting behavior

        self.populateTable()
        self.sorting_order = Qt.AscendingOrder  # Initial sorting order
        # Initialize last_clicked_column attribute
        self.last_clicked_column = -1
        # self.fixFirstColumn()

    def populateTable(self):
        self.horizontalHeader().sectionClicked.connect(self.on_header_clicked)

    def on_header_clicked(self, logical_index):
        first_column = []

        # Iterate through rows to collect data
        for row in range(self.rowCount()):
            item = self.item(row, 0)
            if item is None:  # Handle empty cell case
                text = ""
            else:
                text = item.text()
            row_data = [self.item(row, col).text() if self.item(row, col) else "" for col in
                        range(1, self.columnCount())]
            first_column.append((text, row_data))

        # Sort first_column based on the logical_index
        first_column.sort(key=lambda x: x[1][logical_index - 1] if logical_index > 0 else x[0])

        # Update the table with sorted data
        for row, (text, row_data) in enumerate(first_column):
            self.setItem(row, 0, QTableWidgetItem(text))
            for col, data in enumerate(row_data):
                self.setItem(row, col + 1, QTableWidgetItem(data))

    def fixFirstColumn(self):
        self.verticalHeader().setVisible(False)  # Hide default vertical header

        # Create a new vertical header and set the labels
        new_vertical_header = QHeaderView(Qt.Orientation.Vertical)
        self.setVerticalHeader(new_vertical_header)

        # Set the labels for the first column (virtual vertical header)
        for row in range(self.rowCount()):
            item = QTableWidgetItem(f'Row {row + 1}')
            self.setVerticalHeaderItem(row, item)

        # Adjust sizes to ensure the first column is fixed width
        self.resizeColumnsToContents()
        self.setColumnWidth(0, self.verticalHeader().width())

        # Disable sorting for the first column (if needed)
        self.verticalHeader().setSectionsClickable(False)
        self.verticalHeader().setSortIndicatorShown(False)


class BlinkingLabel(QLabel):
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self._timer = QTimer(self)
        self._timer.timeout.connect(self.toggle_visibility)
        # self._timer.start(1000)
        self.base_text = ""
        self.dot_count = 0
        self.direction = 1

    def start_blinking(self):
        self._timer.start(1000)

    def stop_blinking(self):
        color = QColor(0, 0, 0)
        self.setStyleSheet("QLabel { color: %s }" % color.name())
        self._timer.stop()

    def toggle_visibility(self):
        self.base_text = self.text().replace(".", "")
        self.dot_count += self.direction

        # 如果点的数量到达了5或者到达了0，改变方向
        if self.dot_count == 5:
            self.direction = -1
        elif self.dot_count == 0:
            self.direction = 1

        dots = '.' * self.dot_count
        self.setText(self.base_text + dots)

        color = QColor(random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
        self.setStyleSheet("QLabel { color: %s }" % color.name())
        # self.setVisible(not self.isVisible())


class EventHandler:
    """EventHandler"""

    def __init__(self, parent):
        self.parent = parent

    def button_svn_click(self):
        """SVN開く"""
        try:
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            if self.parent.input_svn.text() is not None:
                open_path = self.parent.input_svn.text()
            folder_path = QFileDialog.getExistingDirectory(self.parent, "SVNパス選択",
                                                           directory=open_path,
                                                           options=options)
            if folder_path:
                print(folder_path)
                self.parent.input_svn.setText(folder_path)
        except Exception as e:
            print("An error occurred : ", e)
            raise

    def button_old_click(self):
        """元ソース開く"""
        try:
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            if self.parent.input_old.text() is not None:
                open_path = self.parent.input_old.text()
            folder_path = QFileDialog.getExistingDirectory(self.parent, "元ソースパス選択",
                                                           directory=open_path,
                                                           options=options)
            if folder_path:
                print(folder_path)
                self.parent.input_old.setText(folder_path)
        except Exception as e:
            print("An error occurred : ", e)
            raise

    def button_new_click(self):
        """新ソース開く"""
        try:
            options = QFileDialog.Options()
            options |= QFileDialog.DontUseNativeDialog
            if self.parent.input_new.text() is not None:
                open_path = self.parent.input_new.text()
            folder_path = QFileDialog.getExistingDirectory(self.parent, "新ソースパス選択",
                                                           directory=open_path,
                                                           options=options)
            if folder_path:
                print(folder_path)
                self.parent.input_new.setText(folder_path)
        except Exception as e:
            print("An error occurred : ", e)
            raise

    def button_manual_click(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        manual_file, _ = QFileDialog.getOpenFileName(None, "マニュアルを選択",
                                                     load_config_content("Paths").get('input_manual', ''),
                                                     "Excel Files (*.xlsx *.xls)", options=options)
        if os.path.exists(manual_file):
            self.parent.input_manual.setText(manual_file)
            # global input_browse
            # input_browse = evidence_file
            # global report_path
            # report_path = evidence_file.split('.xlsx')[0] + load_config_content("Output").get('output_result', '')

    def button_wbs_click(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        wbs_file, _ = QFileDialog.getOpenFileName(None, "外部WBSを選択",
                                                  load_config_content("Paths").get('input_wbs', ''),
                                                  "Excel Files (*.xlsx *.xls)", options=options)
        if os.path.exists(wbs_file):
            self.parent.input_wbs.setText(wbs_file)

    def app_execute(self):
        global EXCEL_TEST, EXCEL_EVIDENCE, EXCEL_LIST, EXCEL_COMPARE, \
            EXCEL_COVERAGE, EXCEL_REVIEW, EXCEL_CD_CHECKLIST, EXCEL_UT_CHECKLIST, EXCEL_TOTAL_MAP
        """tabsに全部のtableをクリア"""
        clearAllTables(self)
        self.parent.progress_bar.setValue(0)
        check_flag, context = is_null_check(self)
        if check_flag is True:
            set_message_box("WARNING", "非空チェック", context[:len(context) - 1])
            return
        # excel_test_path = find_files(self.parent.input_svn.text(), self.parent.input_id.text(), EXCEL_TEST)
        # if excel_test_path is None:
        #     set_message_box("CRITICAL", "ファイル", "機能ID「" + self.parent.input_id.text()
        #                     + "」の成果物が見つかりませんので、チェックしてください。")
        #     return
        # else:
        #     EXCEL_TEST = excel_test_path
        """タブ①、ドキュメントチェック"""
        self.parent.tabs.setCurrentIndex(0)
        set_status_label(self, "ドキュメントチェック中")
        self.document_check()
        self.parent.progress_bar.setValue(20)

        """タブ②、記入内容チェック"""
        self.parent.tabs.setCurrentIndex(1)
        set_status_label(self, "記入内容チェック中")
        self.content_check()
        self.parent.progress_bar.setValue(40)

        """タブ③、チェック"""

        self.parent.progress_bar.setValue(60)

        """タブ④、チェック"""

        self.parent.progress_bar.setValue(80)

        """タブ⑤、チェック"""

        set_status_label(self, "チェック完了しました。")
        self.parent.status_label.stop_blinking()
        # self.parent.save_button.setDisabled(False)
        # self.parent.exec_button.setDisabled(True)
        self.parent.progress_bar.setValue(100)

    def document_check(self):
        global EXCEL_TEST, EXCEL_EVIDENCE, EXCEL_LIST, EXCEL_COMPARE, \
            EXCEL_COVERAGE, EXCEL_REVIEW, EXCEL_CD_CHECKLIST, EXCEL_UT_CHECKLIST, EXCEL_TOTAL_MAP
        context_list = ["エビデンスファイル"]
        self.row_append(0, context_list)
        for i, excel in enumerate(variables):
            if i == 0:
                """テスト仕様書"""
                excel_test_path = find_files_in_svn(self.parent.input_svn.text(), self.parent.input_id.text(),
                                                    variables[excel], True)
                if excel_test_path is not None:
                    svn_result = svn_check_file(excel_test_path)
                    context_list = [None, variables[excel], svn_result, "〇"]
                    EXCEL_TOTAL_MAP[excel] = excel_test_path
                    self.row_append(0, context_list)
                else:
                    return
            else:
                """その他"""
                excel_evidence_path = find_files_in_svn(self.parent.input_svn.text(),
                                                        self.parent.input_id.text(), variables[excel], False)
                if excel_evidence_path is not None:
                    svn_result = svn_check_file(excel_evidence_path)
                    context_list = [None, variables[excel], svn_result, "〇"]
                    EXCEL_TOTAL_MAP[excel] = excel_evidence_path
                    self.row_append(0, context_list)
                else:
                    context_list = [None, variables[excel], "", "✕"]
                    EXCEL_TOTAL_MAP[excel] = None
                    self.row_append(0, context_list)

    def row_append(self, index, context):
        tab_table = self.parent.tabs.widget(index).layout().itemAt(0).widget()
        current_row_count = tab_table.rowCount()
        tab_table.setRowCount(current_row_count + 1)

        # Fill the new row with data
        for col, line in enumerate(context):
            if line is None:
                line = ""
            # item = QTableWidgetItem(f"New Data ({current_row_count + 1}, {line})")
            item = QTableWidgetItem(f"{line}")
            item.setTextAlignment(0x0004 | 0x0080)
            tab_table.setItem(current_row_count, col, item)
        tab_table.resizeColumnsToContents()
        tab_table.resizeRowsToContents()
        QApplication.processEvents()
        bottomRightItem = tab_table.item(current_row_count, 0)
        tab_table.scrollToItem(bottomRightItem)

    def content_check(self):
        global EXCEL_TEST, EXCEL_EVIDENCE, EXCEL_LIST, EXCEL_COMPARE, \
            EXCEL_COVERAGE, EXCEL_REVIEW, EXCEL_CD_CHECKLIST, EXCEL_UT_CHECKLIST, EXCEL_TOTAL_MAP

        name_list = read_name_excel(self.parent.input_wbs.text(), "WBS")
        for name_content in name_list:
            if name_content[4] == self.parent.input_id.text():
                user_name = name_content[18]
                user_date = name_content[21]
                break

        """表紙チェック"""
        self.row_append(1, ["シート「表紙」"])
        for i, excel in enumerate(variables):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')
                # print("content_check : " + excel)
                if EXCEL_TOTAL_MAP[excel] is not None:
                    """表紙チェック"""
                    self.read_hyoushi_sheet(EXCEL_TOTAL_MAP[excel], "表紙", value, user_name, user_date)
        """⑤作業結果確認チェック"""
        self.row_append(1, ["シート「⑤作業結果確認」"])
        for i, excel in enumerate(EXCEL_TOTAL_MAP):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')
                # print("content_check : " + excel)

                if EXCEL_TOTAL_MAP[excel] is not None:
                    self.read_kekka_sheet(EXCEL_TOTAL_MAP[excel], "⑤作業結果確認", value, user_name, user_date)

        """成果物一覧"""
        self.row_append(1, ["シート「成果物一覧」"])
        for i, excel in enumerate(EXCEL_TOTAL_MAP):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')
                # print("content_check : " + excel)

                if EXCEL_TOTAL_MAP[excel] is not None:
                    self.read_ichiran_sheet(EXCEL_TOTAL_MAP[excel], "成果物一覧", value, user_name, user_date)

        """ソースファイル"""
        self.row_append(1, ["シート「ソースファイル」"])
        for i, excel in enumerate(EXCEL_TOTAL_MAP):
            if excel is not None:
                if excel in variables:
                    value = variables[excel]
                    print(value)
                else:
                    print(f'Variable {excel} not found.')
                # print("content_check : " + excel)

                if EXCEL_TOTAL_MAP[excel] is not None:
                    self.read_source_file_sheet(EXCEL_TOTAL_MAP[excel], "ソースファイル", value, user_name, user_date)

    def read_source_file_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date):
        wb = openpyxl.load_workbook(file_path, data_only=False)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            return

        for row in ws.iter_rows(values_only=False):
            for cell in row:
                file_flag = False
                if cell.column_letter == "B" and isinstance(cell.value, int) is True:
                    file_flag = True
                if file_flag is True:
                    self.hyoushi_name_check(row[6].value, wbs_user, value,
                                            f"作成者「{row[6].row}, {row[6].column_letter}」")
                    break
                else:
                    break

    def read_ichiran_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date):
        wb = openpyxl.load_workbook(file_path, data_only=False)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            return

        for row in ws.iter_rows(values_only=False):
            for cell in row:
                file_flag = False
                if cell.column_letter == "A" and isinstance(cell.value, int) is True:
                    file_flag = True
                if file_flag is True:
                    self.hyoushi_name_check(row[4].value, wbs_user, value,
                                            f"作成者「{row[4].row}, {row[4].column_letter}」")
                    break
                else:
                    break

    def read_kekka_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date):
        wb = openpyxl.load_workbook(file_path, data_only=False)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            return

        for i in range(3):
            func_create_user = ws.cell(row=i + 1 + 7, column=column_letter_to_number("AX"))
            self.hyoushi_name_check(func_create_user.value, wbs_user, value,
                                    f"作成者「{func_create_user.row}, {func_create_user.column_letter}」")

    def read_hyoushi_sheet(self, file_path, sheet_name, value, wbs_user, wbs_date):
        wb = openpyxl.load_workbook(file_path, data_only=False)
        try:
            ws = wb[f'{sheet_name}']
        except KeyError:
            return

        func_id = ws.cell(row=2, column=column_letter_to_number("V"))
        self.hyoushi_name_check(func_id.value, self.parent.input_id.text(), value,
                                f"機能ID「{func_id.row}, {func_id.column_letter}」")
        func_name = ws.cell(row=3, column=column_letter_to_number("V"))
        func_sys = ws.cell(row=2, column=column_letter_to_number("F"))
        self.hyoushi_name_check(func_sys.value, self.parent.input_system.currentText(), value,
                                f"システム「{func_sys.row}, {func_sys.column_letter}」")
        func_type = ws.cell(row=2, column=column_letter_to_number("P"))
        self.hyoushi_name_check(func_type.value, self.parent.input_type.currentText(), value,
                                f"区分「{func_type.row}, {func_type.column_letter}」")
        func_create_user = ws.cell(row=2, column=column_letter_to_number("AI"))
        self.hyoushi_name_check(func_create_user.value, "DHC" + wbs_user, value,
                                f"作成者「{func_create_user.row}, {func_create_user.column_letter}」")
        func_update_user = ws.cell(row=3, column=column_letter_to_number("AI"))
        self.hyoushi_name_check(func_update_user.value, "DHC" + wbs_user, value,
                                f"更新者「{func_update_user.row}, {func_update_user.column_letter}」")
        func_create_date = ws.cell(row=2, column=column_letter_to_number("AP")).value
        wb.close()

    def hyoushi_name_check(self, func_cell, cell, value, context):
        if func_cell != cell:
            context_list = [None, value, context, "✕"]
            self.row_append(1, context_list)

    def app_save(self):
        pass

    def app_exit(self):
        """アプリを退出"""
        msg_box = QMessageBox()
        msg_box.setWindowTitle("ツールメッセージ")
        msg_box.setText("ツールを終了したいですか。")
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msg_box.setDefaultButton(QMessageBox.No)
        msg_box.button(QMessageBox.Yes).setText("はい(&Y)")
        msg_box.button(QMessageBox.No).setText("いいえ(&N)")
        result = msg_box.exec_()
        if result == QMessageBox.Yes:
            save_file_paths(self)
            self.parent.close()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.top_group = QGroupBox("選択")
        self.top_layout_1 = QHBoxLayout()
        self.top_layout_2 = QHBoxLayout()
        self.top_layout_3 = QHBoxLayout()
        self.top_layout_4 = QHBoxLayout()
        self.top_layout_5 = QHBoxLayout()
        self.top_layout_6 = QHBoxLayout()
        self.form_layout = QFormLayout()
        self.top_layout = QVBoxLayout()
        self.label_width = 50
        self.label_id = QLabel('機能ID')
        self.label_id.setFixedWidth(self.label_width)
        self.input_id = QLineEdit()
        self.label_type = QLabel('区分')
        self.label_type.setFixedWidth(self.label_width)
        self.input_type = QComboBox()
        self.input_type.addItems(ENABILITY_TYPE)
        self.input_type.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.label_system = QLabel('システム')
        self.label_system.setFixedWidth(self.label_width)
        self.input_system = QComboBox()
        self.input_system.addItems(ENABILITY_SYSTEM)
        self.input_system.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.label_svn = QLabel('SVNパス')
        self.label_svn.setFixedWidth(self.label_width)
        self.input_svn = QLineEdit()
        self.input_svn.setReadOnly(True)
        self.button_svn = QPushButton('開く')
        # self.button_browse.clicked.connect(self.event_handler.browse_button_click)
        # self.button_browse.released.connect(self.event_handler.browse_button_released)
        self.label_manual = QLabel('マニュアル')
        self.label_manual.setFixedWidth(self.label_width)
        self.input_manual = QLineEdit()
        self.input_manual.setReadOnly(True)
        self.button_manual = QPushButton('開く')
        self.label_old = QLabel('元ソース')
        self.label_old.setFixedWidth(self.label_width)
        self.input_old = QLineEdit()
        self.input_old.setReadOnly(True)
        self.button_old = QPushButton('開く')
        self.label_new = QLabel('新ソース')
        self.label_new.setFixedWidth(self.label_width)
        self.input_new = QLineEdit()
        self.input_new.setReadOnly(True)
        self.button_new = QPushButton('開く')
        self.label_wbs = QLabel('外部WBS')
        self.label_wbs.setFixedWidth(self.label_width)
        self.input_wbs = QLineEdit()
        self.input_wbs.setReadOnly(True)
        self.button_wbs = QPushButton('開く')

        self.top_layout_1.addWidget(self.label_id)
        self.top_layout_1.addWidget(self.input_id)
        self.top_layout_1.addWidget(self.label_type)
        self.top_layout_1.addWidget(self.input_type)
        self.top_layout_1.addWidget(self.label_system)
        self.top_layout_1.addWidget(self.input_system)
        self.top_layout_2.addWidget(self.label_svn)
        self.top_layout_2.addWidget(self.input_svn)
        self.top_layout_2.addWidget(self.button_svn)
        self.top_layout_3.addWidget(self.label_manual)
        self.top_layout_3.addWidget(self.input_manual)
        self.top_layout_3.addWidget(self.button_manual)
        self.top_layout_4.addWidget(self.label_old)
        self.top_layout_4.addWidget(self.input_old)
        self.top_layout_4.addWidget(self.button_old)
        self.top_layout_4.addWidget(self.label_new)
        self.top_layout_4.addWidget(self.input_new)
        self.top_layout_4.addWidget(self.button_new)
        self.top_layout_5.addWidget(self.label_wbs)
        self.top_layout_5.addWidget(self.input_wbs)
        self.top_layout_5.addWidget(self.button_wbs)
        self.top_layout.addLayout(self.top_layout_1)
        self.top_layout.addLayout(self.top_layout_2)
        self.top_layout.addLayout(self.top_layout_3)
        self.top_layout.addLayout(self.top_layout_4)
        self.top_layout.addLayout(self.top_layout_5)
        self.top_layout.addLayout(self.top_layout_6)
        self.top_group.setLayout(self.top_layout)
        # self.hbox_layout_1 = QHBoxLayout()
        # self.hbox_layout_1.addWidget(self.label_id)
        # self.hbox_layout_1.addWidget(self.input_id)
        # self.hbox_layout_1.addWidget(self.label_type)
        # self.hbox_layout_1.addWidget(self.input_type)
        # self.hbox_layout_1.addWidget(self.label_system)
        # self.hbox_layout_1.addWidget(self.input_system)
        # self.form_layout.addRow(self.hbox_layout_1)
        # self.hbox_layout_2 = QHBoxLayout()
        # self.hbox_layout_2.addWidget(self.label_svn)
        # self.hbox_layout_2.addWidget(self.input_svn)
        # self.hbox_layout_2.addWidget(self.button_svn)
        # self.form_layout.addRow(self.hbox_layout_2)
        # self.hbox_layout_3 = QHBoxLayout()
        # self.hbox_layout_3.addWidget(self.label_manual)
        # self.hbox_layout_3.addWidget(self.input_manual)
        # self.hbox_layout_3.addWidget(self.button_manual)
        # self.form_layout.addRow(self.hbox_layout_3)
        # self.hbox_layout_4 = QHBoxLayout()
        # self.hbox_layout_4.addWidget(self.label_old)
        # self.hbox_layout_4.addWidget(self.input_old)
        # self.hbox_layout_4.addWidget(self.button_old)
        # self.hbox_layout_4.addWidget(self.label_new)
        # self.hbox_layout_4.addWidget(self.input_new)
        # self.hbox_layout_4.addWidget(self.button_new)
        # self.form_layout.addRow(self.hbox_layout_4)
        # self.hbox_layout_5 = QHBoxLayout()
        # self.hbox_layout_5.addWidget(self.label_wbs)
        # self.hbox_layout_5.addWidget(self.input_wbs)
        # self.hbox_layout_5.addWidget(self.button_wbs)
        # self.form_layout.addRow(self.hbox_layout_5)
        # self.top_layout.addLayout(self.form_layout)
        # self.top_group.setLayout(self.top_layout)

        self.bottom_right_group = QGroupBox('結果')
        self.bottom_right_layout = QVBoxLayout()
        self.tabs = QTabWidget()
        for i in range(5):
            tab = QWidget()
            tab_layout = QVBoxLayout()
            # table = QTableWidget(5, 3)
            # table = SortableTable()
            table = QTableWidget(0, 4)
            table.setHorizontalHeaderLabels(['区分', 'ファイル', '備考', '状態'])
            # for row in range(5):
            #     for col in range(3):
            #         table.setItem(row, col, QTableWidgetItem(f"Step {i + 1} - Cell ({row + 1}, {col + 1})"))
            table_font = table.horizontalHeader().font()
            table_font.setBold(True)
            table.horizontalHeader().setFont(table_font)
            tab_layout.addWidget(table)
            tab.setLayout(tab_layout)
            self.tabs.addTab(tab, f"Step {i + 1}")
            if i == 0:
                self.tabs.addTab(tab, "ドキュメントチェック")
            if i == 1:
                self.tabs.addTab(tab, "記入内容チェック")
            if i == 2:
                self.tabs.addTab(tab, "ソースチェック（作業中）")
            if i == 3:
                self.tabs.addTab(tab, "手修正チェック（作業中）")
            if i == 4:
                self.tabs.addTab(tab, "カバーチェック（作業中）")

        self.tabs.setTabEnabled(2, False)
        self.tabs.setTabEnabled(3, False)
        self.tabs.setTabEnabled(4, False)

        self.bottom_right_layout.addWidget(self.tabs)
        self.bottom_right_group.setLayout(self.bottom_right_layout)

        self.bottom_layout = QHBoxLayout()
        self.bottom_layout.addWidget(self.bottom_right_group)

        self.button_group = QGroupBox("操作")
        self.button_layout = QHBoxLayout()
        self.exec_button = QPushButton('実行')
        # self.exec_button.setDisabled(True)
        # self.exec_button.setStyleSheet(button_stylesheet)
        # self.exec_button.setStyleSheet("background-color: red")
        self.save_button = QPushButton('結果保存（作業中）')
        self.save_button.setDisabled(True)
        self.exit_button = QPushButton('退出')
        self.exit_button.setStyleSheet("background-color: lightgray")
        # self.exec_button.clicked.connect(self.event_handler.execute)
        # self.save_button.clicked.connect(self.event_handler.records_open)
        # self.exit_button.clicked.connect(self.event_handler.app_exit)
        self.button_layout.addWidget(self.exec_button)
        self.button_layout.addWidget(self.save_button)
        self.button_layout.addWidget(self.exit_button)
        self.button_group.setLayout(self.button_layout)

        self.tips_group = QGroupBox("状態")
        self.tips_layout = QVBoxLayout()
        self.tips_layout_1 = QHBoxLayout()
        self.tips_layout_2 = QHBoxLayout()
        self.status_label = BlinkingLabel('画面初期化')
        self.status_label.start_blinking()
        self.tips_label = QLabel('')
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.tips_layout_1.addWidget(self.status_label)
        self.tips_layout_2.addWidget(self.tips_label)
        self.tips_layout_2.addWidget(self.progress_bar)
        self.tips_layout.addLayout(self.tips_layout_1)
        self.tips_layout.addLayout(self.tips_layout_2)
        self.tips_group.setLayout(self.tips_layout)

        self.main_layout = QVBoxLayout()
        self.main_layout.addWidget(self.top_group, 2)
        self.main_layout.addLayout(self.bottom_layout, 5)
        self.main_layout.addWidget(self.button_group, 1)
        self.main_layout.addWidget(self.tips_group, 1)

        central_widget = QWidget()
        central_widget.setLayout(self.main_layout)
        self.setCentralWidget(central_widget)

        self.setLayout(self.main_layout)

        self.setWindowTitle('BIP-整合性チェック-Ver.1.0-Powered by PyQt5')
        self.setGeometry(400, 200, 1000, 700)
        self.timer_init()
        self.event_handler = EventHandler(self)
        self.init_ui()
        init_config_content()
        if INPUT_ID is not None:
            self.input_id.setText(INPUT_ID)
        if INPUT_TYPE is not None:
            self.input_type.setCurrentText(INPUT_TYPE)
        if INPUT_SYSTEM is not None:
            self.input_system.setCurrentText(INPUT_SYSTEM)
        if INPUT_SVN is not None:
            self.input_svn.setText(INPUT_SVN)
        if INPUT_MANUAL is not None:
            self.input_manual.setText(INPUT_MANUAL)
        if INPUT_OLD is not None:
            self.input_old.setText(INPUT_OLD)
        if INPUT_NEW is not None:
            self.input_new.setText(INPUT_NEW)
        if INPUT_WBS is not None:
            self.input_wbs.setText(INPUT_WBS)

    def init_ui(self):
        """init_ui"""
        # self.button_svn.installEventFilter(self)
        # self.button_old.installEventFilter(self)
        # self.button_new.installEventFilter(self)

        self.button_svn.clicked.connect(self.event_handler.button_svn_click)
        self.button_old.clicked.connect(self.event_handler.button_old_click)
        self.button_new.clicked.connect(self.event_handler.button_new_click)
        self.button_manual.clicked.connect(self.event_handler.button_manual_click)
        self.button_wbs.clicked.connect(self.event_handler.button_wbs_click)

        self.exec_button.clicked.connect(self.event_handler.app_execute)
        self.save_button.clicked.connect(self.event_handler.app_save)
        self.exit_button.clicked.connect(self.event_handler.app_exit)

    def closeEvent(self, event):
        config = ConfigParser()
        config_path = get_config_file_path()
        if os.path.exists(config_path) is False:
            return
        config.read(config_path, encoding='utf-8')
        if self.input_id.text():
            config.set('Ids', 'input_id', self.input_id.text())
        if self.input_type.currentText():
            config.set('Ids', 'input_type', self.input_type.currentText())
        if self.input_system.currentText():
            config.set('Ids', 'input_system', self.input_system.currentText())
        if self.input_svn.text():
            config.set('Paths', 'input_svn', self.input_svn.text())
        if self.input_manual.text():
            config.set('Paths', 'input_manual', self.input_manual.text())
        if self.input_old.text():
            config.set('Paths', 'input_old', self.input_old.text())
        if self.input_new.text():
            config.set('Paths', 'input_new', self.input_new.text())
        if self.input_wbs.text():
            config.set('Paths', 'input_wbs', self.input_wbs.text())
        with open(get_config_file_path(), 'w', encoding='utf-8') as configfile:
            config.write(configfile)

    # def eventFilter(self, obj, event):
    #     # if obj == self.input_kinoid and event.type() == event.FocusOut:
    #     #     self.change_text_color(self.input_kinoid.text())
    #     if obj == self.button_svn:
    #         if event.type() == event.MouseButtonRelease:
    #             self.folder_open("SVNパス選択")
    #
    #     if obj == self.button_old:
    #         if event.type() == event.MouseButtonRelease:
    #             self.folder_open("現行ソースパス選択")
    #
    #     if obj == self.button_new:
    #         if event.type() == event.MouseButtonRelease:
    #             self.folder_open("新ソースパス選択")
    #
    #     return super().eventFilter(obj, event)
    #
    # def folder_open(self, folder_name):
    #     """目標開く"""
    #     try:
    #         options = QFileDialog.Options()
    #         options |= QFileDialog.DontUseNativeDialog
    #         if self.input_svn.text() is not None:
    #             open_path = self.input_svn.text()
    #         folder_path = QFileDialog.getExistingDirectory(self, folder_name,
    #                                                        directory=open_path,
    #                                                        options=options)
    #         if folder_path:
    #             print(folder_path)
    #             self.input_svn.setText(folder_path)
    #         return True
    #     except Exception as e:
    #         print("An error occurred : ", e)
    #         raise

    def on_button_clicked(self):
        self.step_bar.advance_step()

    def tab_switched(self, index):
        self.tabs.setCurrentIndex(index)

    def timer_init(self):
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_datetime)
        self.timer.start(1000)

    def update_datetime(self):
        self.current_datetime = QDateTime.currentDateTime().toString(Qt.ISODate)
        self.tips_label.setText(f'{self.current_datetime}')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Windows')  # Windows , windowsvista , Fusion
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
